from openpyxl import load_workbook
import numpy as np
from pyDOE import lhs
import copy
import itertools

def table_extract(excel_file,sheet_name_1,start_cell_1,end_cell_1):
    
    # sheet_name_1 = "自动建模_导出xlsx"
    # start_cell_1 = "B2"
    # end_cell_1 = "C200"
    try:
        # 2. 加载工作簿
        wb = load_workbook(excel_file, data_only=True)
        
        # 3. 获取工作表
        if sheet_name_1 not in wb.sheetnames:
            print(f"错误: 工作簿中不存在工作表 '{sheet_name_1}'")
            print(f"可选工作表: {', '.join(wb.sheetnames)}")
            exit(1)
        
        sheet1 = wb[sheet_name_1]
        # 4. 解析起始和结束单元格的行列号
        # 将"B2"转换为行号2，列号2（B是第2列）
        start_col_1 = ord(start_cell_1[0]) - 64  # 'A'=65, 所以65-64=1
        start_row_1 = int(start_cell_1[1:])
        
        end_col_1 = ord(end_cell_1[0]) - 64
        end_row_1 = int(end_cell_1[1:])

        # 5. 提取数据到二维数组
        data_table_1 = []

        # 使用iter_rows正确遍历区域
        for row in sheet1.iter_rows(min_row=start_row_1, max_row=end_row_1,
                                min_col=start_col_1, max_col=end_col_1,
                                values_only=True):
            row_data_1 = []
            for cell in row:
                value = cell
                if value is None:
                    value = ""
                elif value == "": 
                    value = "<显式空白>"
                elif isinstance(value, str) and value.strip() == "": 
                    value = "<全空格>"
                row_data_1.append(value)
                
            data_table_1.append(row_data_1)

        # # 打印读取的数据
        print(f"成功读取 {len(data_table_1)} 行数据:")
        # for i, row in enumerate(data_table):
        #     print(f"行 {i+1}: {row}")

    except FileNotFoundError:
        print(f"错误: 文件 '{excel_file}' 未找到")
        print("请检查文件路径是否正确，并确保文件存在")
    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")
        print("请检查: 1) 文件格式 2) 是否被其他程序占用 3) 文件内容")

    finally:
        if 'wb' in locals():
            wb.close()
    
    return data_table_1

def latin_hypercube_sampling(num_samples, bounds):
    """
    生成拉丁超立方采样样本
    
    参数:
    num_samples -- 需要生成的样本数量
    bounds -- 包含每个变量上下限的列表，格式为 [(min1, max1), (min2, max2), ...]
    
    返回:
    samples -- 生成的样本数组，形状为 (num_samples, num_variables)
    """
    # 获取变量数量
    num_variables = len(bounds)
    
    # 使用pyDOE生成基础拉丁超立方设计（0-1区间）
    lhs_design = lhs(num_variables, samples=num_samples, criterion='maximin')
    
    # 将设计从0-1区间映射到实际范围
    samples = np.zeros_like(lhs_design)
    for i in range(num_variables):
        min_val, max_val = bounds[i]
        samples[:, i] = lhs_design[:, i] * (max_val - min_val) + min_val
        samples[:, i] = np.round(lhs_design[:, i] * (max_val - min_val) + min_val)
    
    return samples

# 示例用法
if __name__ == "__main__":
    # # 定义采样参数
    # num_samples = 5  # 样本数量
    # bounds = [(0, 10), (20, 50), (-5, 5)]  # 三个变量的范围
    
    # # 生成样本
    # samples = latin_hypercube_sampling(num_samples, bounds)
    excel_file = r"C:\Users\pc\Downloads\油箱建模算单.xlsx"
    data_table = table_extract(excel_file,"自动建模_导出xlsx","B2","C200")
    # data_table = table_extract(excel_file,"采样列表","C2","E5")
    print(data_table)

    # # 打印结果
    # print("拉丁超立方采样结果:")
    # for i, sample in enumerate(samples):
    #     print(f"样本{i+1}: {sample}")

